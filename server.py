from __future__ import annotations

import json
import mimetypes
import os
import posixpath
import re
import time as time_module
import urllib.error
import urllib.request
import cgi
import io
import zipfile
from datetime import date, datetime, time, timedelta
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel


BASE_DIR = Path(__file__).resolve().parent


def load_dotenv() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_dotenv()

SM_DIR = BASE_DIR / "SM"
MASTER_EXCEL = SM_DIR / "汇总表-市场部市场活动计划汇总.xlsx"
HOST = "127.0.0.1"
PORT = int(os.environ.get("MEETING_DASHBOARD_PORT", "8765"))
LOG_FILE = BASE_DIR / "server-runtime.log"
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_ANON_KEY = os.environ.get("SUPABASE_ANON_KEY", "")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_BUCKET = os.environ.get("SUPABASE_BUCKET", "meeting-files")
SUPABASE_MASTER_EXCEL_PATH = os.environ.get("SUPABASE_MASTER_EXCEL_PATH", "data/汇总表-市场部市场活动计划汇总.xlsx")
MEETING_DATA_SOURCE = os.environ.get("MEETING_DATA_SOURCE", "auto").lower()

MASTER_HEADERS = [
    "序号",
    "项目编号",
    "类型",
    "项目类型",
    "项目名称",
    "区域",
    "活动/会议名称",
    "日期",
    "地理位置",
    "项目负责人",
    "参与环节",
    "讲题/内容",
    "参与嘉宾",
    "项目状态",
    "关键进度/问题",
    "备注",
]

EVENT_FIELD_TO_HEADER = {
    "serial": "序号",
    "projectId": "项目编号",
    "type": "类型",
    "projectType": "项目类型",
    "projectName": "项目名称",
    "region": "区域",
    "title": "活动/会议名称",
    "date": "日期",
    "location": "地理位置",
    "owner": "项目负责人",
    "participation": "参与环节",
    "topic": "讲题/内容",
    "guestText": "参与嘉宾",
    "status": "项目状态",
    "progress": "关键进度/问题",
    "notes": "备注",
}

LOCAL_RELEVANT_EXTENSIONS = {
    ".doc",
    ".docx",
    ".pdf",
    ".ppt",
    ".pptx",
    ".xls",
    ".xlsx",
    ".jpg",
    ".jpeg",
    ".png",
}
UPLOAD_EXTENSIONS = LOCAL_RELEVANT_EXTENSIONS
EXCLUDED_ATTACHMENT_KEYWORDS = ("合同", "协议", "汇总表")


def normalize_header(value: Any) -> str:
    return str(value or "").strip().replace("\n", "")


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def excel_date_to_iso(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value).date().isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    match = re.search(r"(20\d{2})[.\-/年]?\s*(\d{1,2})[.\-/月]?\s*(\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return date(year, month, day).isoformat()
    return text


def infer_date_from_text(text: str) -> str:
    match = re.search(r"(20\d{2})[.\-/年]?\s*(\d{1,2})[.\-/月]?\s*(\d{1,2})", text or "")
    if not match:
        return ""
    year, month, day = map(int, match.groups())
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return ""


def date_key(iso_date: str) -> str:
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", iso_date or "")
    return "".join(match.groups()) if match else ""


def split_people(value: str) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[、,，;/；\s]+", value)
    return [part for part in (p.strip() for p in parts) if part]


def safe_relative_path(path: Path) -> str:
    relative = path.resolve().relative_to(SM_DIR.resolve())
    return relative.as_posix()


def file_url(path: Path) -> str:
    return "/files/" + quote(safe_relative_path(path), safe="/")


def is_relevant_attachment(item: Path, category: str) -> bool:
    if item.suffix.lower() not in LOCAL_RELEVANT_EXTENSIONS:
        return False
    searchable = f"{category} {item.name}"
    return not any(keyword in searchable for keyword in EXCLUDED_ATTACHMENT_KEYWORDS)


def normalize_attachment_category(category: str) -> str:
    if "课件" in category or "PPT" in category.upper():
        return "会议课件"
    if "导入" in category or "附件" in category:
        return "其他附件"
    return "其他附件"


def supported_file_label() -> str:
    return "PDF、Word、PPT、Excel、JPG、PNG"


def folder_mtime(path: Path) -> float:
    latest = path.stat().st_mtime if path.exists() else 0
    if path.exists():
        for item in path.rglob("*"):
            try:
                latest = max(latest, item.stat().st_mtime)
            except OSError:
                continue
    return latest


def scan_folders() -> list[Path]:
    if not SM_DIR.exists():
        return []
    return [item for item in SM_DIR.iterdir() if item.is_dir()]


def score_folder(folder: Path, event_date_key: str, title: str) -> int:
    name = folder.name
    score = 0
    if event_date_key and event_date_key in name:
        score += 100
    title_tokens = [token for token in re.split(r"[\s\-（）()_/]+", title) if len(token) >= 2]
    for token in title_tokens:
        if token in name:
            score += min(len(token), 12)
    return score


def match_folder(event_date_key: str, title: str, folders: list[Path]) -> Path | None:
    candidates = sorted(((score_folder(folder, event_date_key, title), folder) for folder in folders), reverse=True)
    if candidates and candidates[0][0] >= 20:
        return candidates[0][1]
    if len(candidates) == 1 and candidates[0][0] >= 8:
        return candidates[0][1]
    if len(candidates) > 1 and candidates[0][0] >= 8 and candidates[0][0] >= candidates[1][0] + 4:
        return candidates[0][1]
    return None


def scan_attachments(folder: Path | None) -> list[dict[str, Any]]:
    if folder is None or not folder.exists():
        return []
    files: list[dict[str, Any]] = []
    for item in folder.rglob("*"):
        if not item.is_file():
            continue
        if item.name.startswith("~$"):
            continue
        try:
            relative_parts = item.relative_to(folder).parts
            category = relative_parts[0] if len(relative_parts) > 1 else "根目录"
            if not is_relevant_attachment(item, category):
                continue
            files.append(
                {
                    "name": item.name,
                    "category": normalize_attachment_category(category),
                    "relativePath": safe_relative_path(item),
                    "url": file_url(item),
                    "extension": item.suffix.lower(),
                    "size": item.stat().st_size,
                    "modifiedAt": datetime.fromtimestamp(item.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                }
            )
        except OSError:
            continue
    return sorted(files, key=lambda f: (f["category"], f["name"]))


def extract_local_excel_tables(folder: Path | None) -> list[dict[str, Any]]:
    if folder is None:
        return []
    tables: list[dict[str, Any]] = []
    for excel_path in folder.rglob("*.xlsx"):
        if excel_path.name.startswith("~$") or excel_path.resolve() == MASTER_EXCEL.resolve():
            continue
        try:
            workbook = load_workbook(excel_path, data_only=True, read_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
            rows = []
            for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 12), values_only=True):
                values = [clean_value(cell) for cell in row]
                if any(values):
                    rows.append(values)
            if rows:
                tables.append(
                    {
                        "fileName": excel_path.name,
                        "sheetName": worksheet.title,
                        "url": file_url(excel_path),
                        "rows": rows,
                    }
                )
            workbook.close()
        except Exception as exc:
            tables.append({"fileName": excel_path.name, "sheetName": "", "url": file_url(excel_path), "error": str(exc)})
    return tables[:3]


def supabase_configured() -> bool:
    return bool(SUPABASE_URL and (SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY) and SUPABASE_BUCKET)


def storage_event_id(value: str) -> str:
    cleaned = re.sub(r"[^\w\u4e00-\u9fff.-]+", "-", value or "unknown", flags=re.UNICODE).strip(".-")
    return cleaned or "unknown"


def clean_storage_filename(value: str) -> str:
    name = Path(value or "attachment").name
    name = re.sub(r"[\r\n\t/\\]+", "-", name).strip(" .")
    return name or f"attachment-{int(time_module.time())}"


def storage_headers(content_type: str = "application/json") -> dict[str, str]:
    key = SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY
    return {
        "Authorization": f"Bearer {key}",
        "apikey": key,
        "Content-Type": content_type,
    }


def storage_request(
    method: str,
    url: str,
    body: bytes | None = None,
    headers: dict[str, str] | None = None,
) -> tuple[int, dict[str, str], bytes]:
    request = urllib.request.Request(url, data=body, headers=headers or {}, method=method)
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return response.status, dict(response.headers.items()), response.read()
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(detail or str(exc)) from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(str(exc)) from exc


def storage_object_url(path: str) -> str:
    return f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{quote(path, safe='/')}"


def get_storage_object(path: str) -> bytes:
    _, _, body = storage_request("GET", storage_object_url(path), None, storage_headers("application/octet-stream"))
    return body


def put_storage_object(path: str, data: bytes, content_type: str) -> None:
    headers = storage_headers(content_type)
    headers["x-upsert"] = "true"
    storage_request("POST", storage_object_url(path), data, headers)


def use_supabase_master() -> bool:
    if MEETING_DATA_SOURCE == "supabase":
        return True
    if MEETING_DATA_SOURCE == "local":
        return False
    return not MASTER_EXCEL.exists() and supabase_configured()


def create_empty_master_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "会议汇总"
    worksheet.append(MASTER_HEADERS)
    format_master_sheet(worksheet)
    return workbook


def load_master_workbook(data_only: bool = True, read_only: bool = True):
    if use_supabase_master():
        if not supabase_configured():
            raise FileNotFoundError("未找到本地总表，且未配置 Supabase Storage 主数据源")
        try:
            data = get_storage_object(SUPABASE_MASTER_EXCEL_PATH)
            return load_workbook(io.BytesIO(data), data_only=data_only, read_only=read_only)
        except Exception:
            if read_only:
                raise
            return create_empty_master_workbook()
    if not MASTER_EXCEL.exists():
        raise FileNotFoundError(f"未找到总表：{MASTER_EXCEL}")
    return load_workbook(MASTER_EXCEL, data_only=data_only, read_only=read_only)


def save_master_workbook(workbook) -> None:
    if use_supabase_master():
        buffer = io.BytesIO()
        workbook.save(buffer)
        put_storage_object(
            SUPABASE_MASTER_EXCEL_PATH,
            buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        return
    MASTER_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(MASTER_EXCEL)


def list_cloud_files(event_id: str) -> list[dict[str, Any]]:
    if not supabase_configured():
        return []
    prefix = f"{storage_event_id(event_id)}/"
    payload = json.dumps(
        {
            "prefix": prefix,
            "limit": 100,
            "offset": 0,
            "sortBy": {"column": "name", "order": "asc"},
        }
    ).encode("utf-8")
    url = f"{SUPABASE_URL}/storage/v1/object/list/{SUPABASE_BUCKET}"
    _, _, body = storage_request("POST", url, payload, storage_headers())
    items = json.loads(body.decode("utf-8") or "[]")
    files: list[dict[str, Any]] = []
    for item in items:
        name = item.get("name", "")
        if not name or name.endswith("/"):
            continue
        path = name if name.startswith(prefix) else f"{prefix}{name}"
        basename = path.removeprefix(prefix)
        files.append(
            {
                "name": basename,
                "path": path,
                "url": "/api/cloud-download?path=" + quote(path, safe="/"),
                "modifiedAt": clean_value(item.get("updated_at") or item.get("created_at"))[:16],
                "size": item.get("metadata", {}).get("size", 0),
            }
        )
    return files


def upload_cloud_file(event_id: str, filename: str, content_type: str, data: bytes) -> dict[str, Any]:
    if not supabase_configured():
        raise RuntimeError("未配置 Supabase Storage")
    safe_name = clean_storage_filename(filename)
    extension = Path(safe_name).suffix.lower()
    if extension not in UPLOAD_EXTENSIONS:
        raise RuntimeError(f"仅支持上传 {supported_file_label()} 文件")
    path = f"{storage_event_id(event_id)}/{safe_name}"
    headers = storage_headers(content_type or mimetypes.guess_type(safe_name)[0] or "application/octet-stream")
    headers["x-upsert"] = "true"
    storage_request("POST", storage_object_url(path), data, headers)
    return {
        "name": safe_name,
        "path": path,
        "url": "/api/cloud-download?path=" + quote(path, safe="/"),
    }


def find_event_folder(event_id: str) -> Path | None:
    events_payload = read_master_events()
    for event in events_payload.get("events", []):
        if event.get("id") != event_id:
            continue
        folder_path = event.get("folderPath")
        if not folder_path:
            return None
        folder = (SM_DIR / Path(folder_path)).resolve()
        try:
            folder.relative_to(SM_DIR.resolve())
        except ValueError:
            return None
        return folder if folder.exists() and folder.is_dir() else None
    return None


def import_local_file(event_id: str, filename: str, data: bytes) -> dict[str, Any]:
    folder = find_event_folder(event_id)
    if folder is None:
        raise RuntimeError("未找到该会议的资料文件夹，无法导入文件")
    safe_name = clean_storage_filename(filename)
    extension = Path(safe_name).suffix.lower()
    if extension not in UPLOAD_EXTENSIONS:
        raise RuntimeError(f"仅支持导入 {supported_file_label()} 文件")
    target_dir = folder / "导入附件"
    target_dir.mkdir(exist_ok=True)
    target = (target_dir / safe_name).resolve()
    target.relative_to(SM_DIR.resolve())
    if target.exists():
        stem = target.stem
        suffix = target.suffix
        target = target.with_name(f"{stem}-{datetime.now().strftime('%Y%m%d%H%M%S')}{suffix}")
    target.write_bytes(data)
    return {
        "name": target.name,
        "category": "其他附件",
        "relativePath": safe_relative_path(target),
        "url": file_url(target),
        "extension": target.suffix.lower(),
        "size": target.stat().st_size,
        "modifiedAt": datetime.fromtimestamp(target.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
    }


def worksheet_headers(worksheet) -> list[str]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    return [normalize_header(cell) for cell in header_row]


def ensure_master_headers(worksheet) -> list[str]:
    headers = worksheet_headers(worksheet)
    if not any(headers):
        worksheet.append(MASTER_HEADERS)
        return MASTER_HEADERS[:]
    for header in MASTER_HEADERS:
        if header not in headers:
            worksheet.cell(row=1, column=len(headers) + 1, value=header)
            headers.append(header)
    return headers


def format_master_sheet(worksheet) -> None:
    headers = ensure_master_headers(worksheet)
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="17324D")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "序号": 8,
        "项目编号": 16,
        "类型": 12,
        "项目类型": 14,
        "项目名称": 18,
        "区域": 12,
        "活动/会议名称": 34,
        "日期": 14,
        "地理位置": 28,
        "项目负责人": 14,
        "参与环节": 16,
        "讲题/内容": 22,
        "参与嘉宾": 24,
        "项目状态": 14,
        "关键进度/问题": 30,
        "备注": 20,
    }
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = widths.get(header, 16)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def next_serial(worksheet, headers: list[str]) -> int:
    if "序号" not in headers:
        return max(1, worksheet.max_row)
    serial_col = headers.index("序号") + 1
    values = []
    for row in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=serial_col).value
        try:
            values.append(int(value))
        except (TypeError, ValueError):
            continue
    return (max(values) + 1) if values else 1


def find_event_row(worksheet, headers: list[str], event_id: str) -> int | None:
    if "项目编号" not in headers:
        return None
    project_col = headers.index("项目编号") + 1
    for row in range(2, worksheet.max_row + 1):
        if clean_value(worksheet.cell(row=row, column=project_col).value) == event_id:
            return row
    return None


def generate_project_id(event: dict[str, Any], worksheet, headers: list[str]) -> str:
    event_date = clean_value(event.get("date")) or datetime.now().date().isoformat()
    key = date_key(excel_date_to_iso(event_date)) or datetime.now().strftime("%Y%m%d")
    base = f"GE{key}"
    existing = set()
    if "项目编号" in headers:
        project_col = headers.index("项目编号") + 1
        existing = {clean_value(worksheet.cell(row=row, column=project_col).value) for row in range(2, worksheet.max_row + 1)}
    index = 1
    candidate = f"{base}{index:02d}"
    while candidate in existing:
        index += 1
        candidate = f"{base}{index:02d}"
    return candidate


def normalize_event_payload(payload: dict[str, Any]) -> dict[str, str]:
    normalized = {key: clean_value(payload.get(key)) for key in EVENT_FIELD_TO_HEADER}
    normalized["projectId"] = normalized.get("projectId") or clean_value(payload.get("id"))
    normalized["title"] = normalized.get("title") or clean_value(payload.get("活动/会议名称"))
    normalized["date"] = excel_date_to_iso(normalized.get("date") or payload.get("日期"))
    normalized["guestText"] = normalized.get("guestText") or clean_value(payload.get("guests"))
    return normalized


def write_event_to_row(worksheet, headers: list[str], row: int, event: dict[str, str], serial: int | None = None) -> str:
    if not event.get("projectId"):
        event["projectId"] = generate_project_id(event, worksheet, headers)
    if serial is not None:
        event["serial"] = str(serial)
    for field, header in EVENT_FIELD_TO_HEADER.items():
        if header not in headers:
            continue
        col = headers.index(header) + 1
        worksheet.cell(row=row, column=col, value=event.get(field, ""))
    return event["projectId"]


def add_master_event(payload: dict[str, Any]) -> dict[str, Any]:
    event = normalize_event_payload(payload)
    if not event.get("title"):
        raise RuntimeError("请填写会议名称")
    workbook = load_master_workbook(data_only=False, read_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = ensure_master_headers(worksheet)
    project_id = event.get("projectId")
    if project_id and find_event_row(worksheet, headers, project_id):
        raise RuntimeError(f"项目编号已存在：{project_id}")
    row = worksheet.max_row + 1
    event["projectId"] = project_id or generate_project_id(event, worksheet, headers)
    write_event_to_row(worksheet, headers, row, event, next_serial(worksheet, headers))
    format_master_sheet(worksheet)
    save_master_workbook(workbook)
    workbook.close()
    return {"projectId": event["projectId"]}


def delete_master_event(event_id: str) -> None:
    workbook = load_master_workbook(data_only=False, read_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = ensure_master_headers(worksheet)
    row = find_event_row(worksheet, headers, event_id)
    if row is None:
        workbook.close()
        raise RuntimeError("未找到要删除的会议")
    worksheet.delete_rows(row, 1)
    format_master_sheet(worksheet)
    save_master_workbook(workbook)
    workbook.close()


def export_events_workbook() -> bytes:
    payload = read_master_events()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "会议汇总"
    worksheet.append(MASTER_HEADERS)
    for index, event in enumerate(payload.get("events", []), start=1):
        row_values = []
        for header in MASTER_HEADERS:
            field = next((field for field, mapped in EVENT_FIELD_TO_HEADER.items() if mapped == header), "")
            row_values.append(index if header == "序号" else event.get(field, ""))
        worksheet.append(row_values)
    format_master_sheet(worksheet)
    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def import_events_workbook(data: bytes) -> dict[str, int]:
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    source = workbook[workbook.sheetnames[0]]
    source_headers = worksheet_headers(source)
    if not any(header in source_headers for header in ("项目编号", "活动/会议名称", "日期")):
        workbook.close()
        raise RuntimeError("导入文件缺少标准表头，请先导出模板后填写")

    master = load_master_workbook(data_only=False, read_only=False)
    worksheet = master[master.sheetnames[0]]
    headers = ensure_master_headers(worksheet)
    created = 0
    updated = 0
    skipped = 0

    for source_row in source.iter_rows(min_row=2, values_only=True):
        values = {source_headers[index]: source_row[index] if index < len(source_headers) else None for index in range(len(source_headers))}
        if not any(clean_value(value) for value in values.values()):
            continue
        payload = {}
        for field, header in EVENT_FIELD_TO_HEADER.items():
            payload[field] = values.get(header)
        event = normalize_event_payload(payload)
        if not event.get("title") and not event.get("date"):
            skipped += 1
            continue
        row = find_event_row(worksheet, headers, event.get("projectId", "")) if event.get("projectId") else None
        if row:
            write_event_to_row(worksheet, headers, row, event)
            updated += 1
        else:
            row = worksheet.max_row + 1
            if not event.get("projectId"):
                event["projectId"] = generate_project_id(event, worksheet, headers)
            write_event_to_row(worksheet, headers, row, event, next_serial(worksheet, headers))
            created += 1
    workbook.close()
    format_master_sheet(worksheet)
    save_master_workbook(master)
    master.close()
    return {"created": created, "updated": updated, "skipped": skipped}


def read_master_events() -> dict[str, Any]:
    workbook = load_master_workbook(data_only=True, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = worksheet_headers(worksheet)
    folders = scan_folders()
    events: list[dict[str, Any]] = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        title = clean_value(values.get("活动/会议名称")) or clean_value(values.get("项目名称")) or f"第 {row_index} 行会议"
        event_date = excel_date_to_iso(values.get("日期")) or infer_date_from_text(title)
        event_date_key = date_key(event_date)
        folder = match_folder(event_date_key, title, folders)
        attachments = scan_attachments(folder)
        local_tables = extract_local_excel_tables(folder)
        project_id = clean_value(values.get("项目编号")) or f"row-{row_index}"

        events.append(
            {
                "id": project_id,
                "rowIndex": row_index,
                "serial": clean_value(values.get("序号")),
                "projectId": project_id,
                "type": clean_value(values.get("类型")),
                "projectType": clean_value(values.get("项目类型")),
                "projectName": clean_value(values.get("项目名称")),
                "title": title,
                "region": clean_value(values.get("区域")),
                "date": event_date,
                "dateKey": event_date_key,
                "location": clean_value(values.get("地理位置")),
                "owner": clean_value(values.get("项目负责人")),
                "participation": clean_value(values.get("参与环节")),
                "topic": clean_value(values.get("讲题/内容")),
                "guests": split_people(clean_value(values.get("参与嘉宾"))),
                "guestText": clean_value(values.get("参与嘉宾")),
                "status": clean_value(values.get("项目状态")),
                "progress": clean_value(values.get("关键进度/问题")),
                "notes": clean_value(values.get("备注")),
                "folderName": folder.name if folder else "",
                "folderPath": safe_relative_path(folder) if folder else "",
                "folderMatched": bool(folder),
                "attachments": attachments,
                "localTables": local_tables,
            }
        )

    workbook.close()
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    local_source_mtime = MASTER_EXCEL.stat().st_mtime if MASTER_EXCEL.exists() else datetime.now().timestamp()
    mtimes = [local_source_mtime, *(folder_mtime(folder) for folder in folders)]
    latest_mtime = max(mtimes)
    return {
        "generatedAt": generated_at,
        "sourceFile": MASTER_EXCEL.name if not use_supabase_master() else SUPABASE_MASTER_EXCEL_PATH,
        "sourceModifiedAt": datetime.fromtimestamp(local_source_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "dataVersion": latest_mtime,
        "count": len(events),
        "events": events,
    }


class MeetingDashboardHandler(SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def send_json(self, payload: dict[str, Any], status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json_body(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0") or 0)
        if length <= 0:
            return {}
        data = self.rfile.read(length)
        return json.loads(data.decode("utf-8") or "{}")

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/events":
            self.serve_events()
            return
        if parsed.path == "/api/cloud-files":
            event_id = parse_qs(parsed.query).get("eventId", [""])[0]
            self.serve_cloud_files(event_id)
            return
        if parsed.path == "/api/cloud-download":
            path = parse_qs(parsed.query).get("path", [""])[0]
            self.serve_cloud_download(path)
            return
        if parsed.path == "/api/export-files":
            event_id = parse_qs(parsed.query).get("eventId", [""])[0]
            self.serve_export_files(event_id)
            return
        if parsed.path == "/api/export-events":
            self.serve_export_events()
            return
        if parsed.path.startswith("/files/"):
            self.serve_sm_file(parsed.path.removeprefix("/files/"))
            return
        if parsed.path == "/":
            self.path = "/index.html"
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/events":
            self.handle_add_event()
            return
        if parsed.path == "/api/upload":
            event_id = parse_qs(parsed.query).get("eventId", [""])[0]
            self.handle_upload(event_id)
            return
        if parsed.path == "/api/import-file":
            event_id = parse_qs(parsed.query).get("eventId", [""])[0]
            self.handle_import_file(event_id)
            return
        if parsed.path == "/api/import-events":
            self.handle_import_events()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "接口不存在")

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/events":
            event_id = parse_qs(parsed.query).get("eventId", [""])[0]
            self.handle_delete_event(event_id)
            return
        self.send_error(HTTPStatus.NOT_FOUND, "接口不存在")

    def serve_events(self) -> None:
        try:
            payload = read_master_events()
            self.send_json(payload)
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.INTERNAL_SERVER_ERROR)

    def handle_add_event(self) -> None:
        try:
            result = add_master_event(self.read_json_body())
            self.send_json({"ok": True, **result})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_delete_event(self, event_id: str) -> None:
        if not event_id:
            self.send_json({"error": "缺少会议 ID"}, HTTPStatus.BAD_REQUEST)
            return
        try:
            delete_master_event(event_id)
            self.send_json({"ok": True})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def serve_cloud_files(self, event_id: str) -> None:
        if not event_id:
            self.send_json({"configured": supabase_configured(), "files": []})
            return
        try:
            self.send_json({"configured": supabase_configured(), "files": list_cloud_files(event_id)})
        except Exception as exc:
            self.send_json({"configured": True, "error": str(exc), "files": []}, HTTPStatus.BAD_GATEWAY)

    def serve_cloud_download(self, path: str) -> None:
        if not supabase_configured():
            self.send_error(HTTPStatus.SERVICE_UNAVAILABLE, "未配置 Supabase Storage")
            return
        path = posixpath.normpath(unquote(path)).lstrip("/")
        if not path or path.startswith("../") or "/../" in path:
            self.send_error(HTTPStatus.FORBIDDEN, "文件路径无效")
            return
        try:
            _, headers, data = storage_request("GET", storage_object_url(path), None, storage_headers("application/octet-stream"))
        except Exception as exc:
            self.send_error(HTTPStatus.BAD_GATEWAY, str(exc))
            return
        filename = Path(path).name
        content_type = headers.get("Content-Type") or mimetypes.guess_type(filename)[0] or "application/octet-stream"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
        self.end_headers()
        self.wfile.write(data)

    def handle_upload(self, event_id: str) -> None:
        if not event_id:
            self.send_json({"error": "缺少会议 ID"}, HTTPStatus.BAD_REQUEST)
            return
        if not supabase_configured():
            self.send_json({"error": "未配置 Supabase Storage"}, HTTPStatus.SERVICE_UNAVAILABLE)
            return
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            self.send_json({"error": "请选择要上传的文件"}, HTTPStatus.BAD_REQUEST)
            return
        data = file_item.file.read()
        if len(data) > 50 * 1024 * 1024:
            self.send_json({"error": "单个文件不能超过 50MB"}, HTTPStatus.BAD_REQUEST)
            return
        content_type = getattr(file_item, "type", "") or mimetypes.guess_type(file_item.filename)[0] or "application/octet-stream"
        try:
            uploaded = upload_cloud_file(event_id, file_item.filename, content_type, data)
            self.send_json({"ok": True, "file": uploaded})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_GATEWAY)

    def handle_import_file(self, event_id: str) -> None:
        if not event_id:
            self.send_json({"error": "缺少会议 ID"}, HTTPStatus.BAD_REQUEST)
            return
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            self.send_json({"error": "请选择要导入的文件"}, HTTPStatus.BAD_REQUEST)
            return
        data = file_item.file.read()
        if len(data) > 80 * 1024 * 1024:
            self.send_json({"error": "单个文件不能超过 80MB"}, HTTPStatus.BAD_REQUEST)
            return
        try:
            imported = import_local_file(event_id, file_item.filename, data)
            self.send_json({"ok": True, "file": imported})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_import_events(self) -> None:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
            },
        )
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            self.send_json({"error": "请选择要导入的会议 Excel"}, HTTPStatus.BAD_REQUEST)
            return
        if Path(file_item.filename).suffix.lower() not in {".xlsx", ".xlsm"}:
            self.send_json({"error": "仅支持导入 .xlsx / .xlsm 会议表"}, HTTPStatus.BAD_REQUEST)
            return
        data = file_item.file.read()
        if len(data) > 20 * 1024 * 1024:
            self.send_json({"error": "会议 Excel 不能超过 20MB"}, HTTPStatus.BAD_REQUEST)
            return
        try:
            result = import_events_workbook(data)
            self.send_json({"ok": True, **result})
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def serve_export_events(self) -> None:
        try:
            data = export_events_workbook()
        except Exception as exc:
            self.send_error(HTTPStatus.INTERNAL_SERVER_ERROR, str(exc))
            return
        filename = f"会议汇总导出-{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
        self.end_headers()
        self.wfile.write(data)

    def serve_export_files(self, event_id: str) -> None:
        if not event_id:
            self.send_error(HTTPStatus.BAD_REQUEST, "缺少会议 ID")
            return
        folder = find_event_folder(event_id)
        if folder is None:
            self.send_error(HTTPStatus.NOT_FOUND, "未找到该会议的资料文件夹")
            return
        files = scan_attachments(folder)
        if not files:
            self.send_error(HTTPStatus.NOT_FOUND, "当前会议暂无可导出的相关文件")
            return
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file in files:
                source = (SM_DIR / Path(file["relativePath"])).resolve()
                try:
                    source.relative_to(SM_DIR.resolve())
                except ValueError:
                    continue
                if source.is_file():
                    archive.write(source, arcname=f"{file['category']}/{source.name}")
        data = buffer.getvalue()
        filename = f"{storage_event_id(event_id)}-相关文件.zip"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/zip")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(filename)}")
        self.end_headers()
        self.wfile.write(data)

    def serve_sm_file(self, encoded_relative_path: str) -> None:
        relative = unquote(encoded_relative_path)
        relative = posixpath.normpath(relative).lstrip("/")
        target = (SM_DIR / Path(relative)).resolve()
        try:
            target.relative_to(SM_DIR.resolve())
        except ValueError:
            self.send_error(HTTPStatus.FORBIDDEN, "文件访问被限制在 SM 目录内")
            return
        if not target.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "文件不存在")
            return
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        try:
            data = target.read_bytes()
        except OSError:
            self.send_error(HTTPStatus.INTERNAL_SERVER_ERROR, "文件读取失败")
            return
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f"inline; filename*=UTF-8''{quote(target.name)}")
        self.end_headers()
        self.wfile.write(data)


handler = MeetingDashboardHandler


def main() -> None:
    os.chdir(BASE_DIR)
    server = ThreadingHTTPServer((HOST, PORT), MeetingDashboardHandler)
    LOG_FILE.write_text(
        f"会议可视化页面已启动：http://{HOST}:{PORT}\n主数据源：{MASTER_EXCEL}\n",
        encoding="utf-8",
    )
    server.serve_forever()


if __name__ == "__main__":
    main()
